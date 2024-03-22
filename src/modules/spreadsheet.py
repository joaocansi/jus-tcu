from openpyxl import Workbook
from openpyxl.styles import Alignment

from config.spreadsheet import columns, columns_settings
from openpyxl.styles import Font, Alignment

from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from openpyxl.utils import get_column_letter

bold = InlineFont(b=True)
b = TextBlock

class Spreadsheet:
    def __init__(self, reports):
        self.reports = reports
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        
    def populate(self):    
        self.worksheet.append(columns)
        for report in self.reports:
            for judgment in report.judgments:
                self.worksheet.append(self._get_worksheet_judgment(report, judgment))
                
        self._setup()
        self.workbook.save('file.xlsx')
        return
    
    def _setup(self):
        for letter, settings in columns_settings.items():
            column = self.worksheet.column_dimensions[letter]
            column.width = settings.width
        
        for row in self.worksheet[2:self.worksheet.max_row]:
            for cell in row:
                cell.alignment = columns_settings[cell.column_letter].alignment
        
        cell = self.worksheet['A2']
        self.worksheet.freeze_panes = cell
        
        FullRange = "A1:" + get_column_letter(self.worksheet.max_column) + str(self.worksheet.max_row)
        self.worksheet.auto_filter.ref = FullRange
    
    def _get_worksheet_judgment(self, report, judgment):
        content = CellRichText(
            b(bold, judgment.header + '\n'),
            judgment.body
        )
        return (judgment.type, judgment.subject, int(judgment.judgment), judgment.session_date, judgment.rapporteur, judgment.process, int(report.report), content, report.release_date, int(report.release_date.split('/')[2]))